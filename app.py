import streamlit as st
import yfinance as yf
import pandas as pd
import numpy as np
import requests
import io
import openpyxl                 
from fpdf import FPDF
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side 
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas_ta as ta
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# --- THE GRAND UNIFICATION IMPORTS ---
try:
    from dcf_nlp import run_sentiment_analysis
    HAS_NLP = True
except Exception as e:
    HAS_NLP = False
    NLP_ERROR = e

try:
    from dcf_consensus import get_dupont_and_health_metrics
    HAS_CONSENSUS = True
except Exception as e:
    HAS_CONSENSUS = False
    CONSENSUS_ERROR = e

try:
    from dcf_altdata import run_altdata_valuation
    HAS_ALTDATA = True
except Exception as e:
    HAS_ALTDATA = False
    ALTDATA_ERROR = e

st.set_page_config(page_title="Financial Research System", page_icon="💹", layout="wide")

st.markdown("""
<style>
    .stApp { background-color: #0b0f17; color: #e2e8f0; }
    .conclusion-box { background-color: #161b22; border-left: 4px solid #3b82f6; padding: 20px; margin-top: -15px; margin-bottom: 50px; border-radius: 0 0 8px 8px; font-size: 0.95rem; color: #94a3b8; line-height: 1.6; }
    .module-box { background-color: #1e293b; border: 1px solid #334155; border-left: 4px solid #8b5cf6; padding: 20px; border-radius: 8px; margin-bottom: 30px; font-size: 0.95rem; line-height: 1.5; }
    h2 { color: #60a5fa; margin-top: 40px; text-transform: uppercase; font-size: 1.3rem; border-bottom: 1px solid #1e293b; padding-bottom: 10px;}
</style>
""", unsafe_allow_html=True)

def clean_for_pdf(text):
    """
    Cleans HTML tags and special characters so the PDF report 
    looks professional and doesn't show code.
    """
    replacements = {
        "<b>": "", 
        "</b>": "", 
        "<br>": "\n", 
        "🟢": "", 
        "🔴": "", 
        "⚪": "", 
        "•": "-", 
        "[-]": "-", 
        "**": ""
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text
import time
import random
from threading import Lock

# 1. Global Traffic Cop: Forces all users into a single-file line
api_lock = Lock()

def safe_api_call(func, *args, **kwargs):
    """
    Wraps all Yahoo Finance calls in a protective shield.
    Includes mandatory human-like delays and exponential backoff for 429 errors.
    """
    max_retries = 3
    for attempt in range(max_retries):
        with api_lock: # Forces users to wait their turn
            try:
                # Mandatory 1 to 2 second delay so Yahoo doesn't think it's a bot attack
                time.sleep(random.uniform(1.0, 2.0)) 
                return func(*args, **kwargs)
            except Exception as e:
                error_msg = str(e).lower()
                # If Yahoo throws a ban, wait and try again
                if "429" in error_msg or "too many requests" in error_msg or "rate limit" in error_msg:
                    if attempt < max_retries - 1:
                        backoff_time = (2 ** attempt) + random.uniform(0.1, 1.0)
                        time.sleep(backoff_time)
                        continue
                # If it's a real error (like a fake ticker), pass it down
                raise e

def resolve_ticker(user_input):
    user_input = user_input.strip()
    if '.' in user_input: return user_input.upper()
    try:
        url = f"https://query2.finance.yahoo.com/v1/finance/search?q={user_input}"
        headers = {'User-Agent': 'Mozilla/5.0'} 
        response = requests.get(url, headers=headers).json()
        quotes = response.get('quotes', [])
        for quote in quotes:
            if quote.get('exchange') in ['NSI', 'BSE']: return quote['symbol']
        for quote in quotes:
            if quote.get('quoteType') == 'EQUITY': return quote['symbol']
        return user_input.upper()
    except: return user_input.upper()


# ==========================================
# PART 1: THE EXCEL GENERATOR FUNCTION
# ==========================================
def generate_dcf_excel(ticker_symbol, base_revenue, shares_out, total_cash, total_debt, actual_margin, actual_growth, stock_currency, financial_currency, live_fx_rate):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proprietary DCF Model"

    header_fill = PatternFill(start_color="1B2631", end_color="1B2631", fill_type="solid")
    subheader_fill = PatternFill(start_color="EAEDED", end_color="EAEDED", fill_type="solid")
    input_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    blue_font = Font(color="0000FF")
    bold_font = Font(bold=True)
    top_bottom_border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    ws.column_dimensions['A'].width = 38
    for i in range(2, 9): ws.column_dimensions[get_column_letter(i)].width = 16

    ws['A1'] = f"DISCOUNTED CASH FLOW (DCF) MODEL - {ticker_symbol.upper()}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Format: Institutional Standard (Unlevered FCF)"

    headers = ["(Figures in Millions)", "Year 0 (Current)", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.fill = header_fill; cell.font = white_font

    ws['A5'] = "1. UNLEVERED FREE CASH FLOW BUILD"
    ws['A5'].fill = subheader_fill; ws['A5'].font = bold_font

    ws['A6'] = "Total Revenue"
    ws['B6'] = base_revenue
    ws['A7'] = "  % YoY Growth"

    for c in range(3, 8):
        col = get_column_letter(c); prev_col = get_column_letter(c-1)
        ws[f'{col}7'] = actual_growth
        ws[f'{col}7'].font = blue_font; ws[f'{col}7'].fill = input_fill; ws[f'{col}7'].number_format = '0.0%'
        ws[f'{col}6'] = f"={prev_col}6*(1+{col}7)"; ws[f'{col}6'].number_format = '#,##0'

    ws['A8'] = "EBIT (Operating Income)"
    ws['A9'] = "  % EBIT Margin"
    ws['A10'] = "Less: Taxes (25%)"
    ws['A11'] = "EBIAT (Net Operating Profit After Tax)"

    for c in range(2, 8):
        col = get_column_letter(c)
        ws[f'{col}9'] = actual_margin
        ws[f'{col}9'].font = blue_font; ws[f'{col}9'].fill = input_fill; ws[f'{col}9'].number_format = '0.0%'
        ws[f'{col}8'] = f"={col}6*{col}9"
        ws[f'{col}10'] = f"=-{col}8*0.25"
        ws[f'{col}11'] = f"={col}8+{col}10"; ws[f'{col}11'].font = bold_font

    ws['A12'] = "Plus: Depreciation & Amortization"
    ws['A13'] = "Less: Capital Expenditure"
    ws['A14'] = "Less: Change in Net Working Capital"
    ws['A15'] = "UNLEVERED FREE CASH FLOW"

    for c in range(2, 8):
        col = get_column_letter(c)
        ws[f'{col}12'] = f"={col}6*0.04"
        ws[f'{col}13'] = f"=-{col}6*0.05"
        ws[f'{col}14'] = f"=-{col}6*0.01"
        ws[f'{col}15'] = f"=SUM({col}11:{col}14)"; ws[f'{col}15'].font = bold_font; ws[f'{col}15'].border = top_bottom_border

    ws['A17'] = "2. DISCOUNTING & TERMINAL VALUE"
    ws['A17'].fill = subheader_fill; ws['A17'].font = bold_font

    ws['A18'] = "Weighted Average Cost of Capital (WACC)"
    ws['B18'] = 0.10
    ws['B18'].font = blue_font; ws['B18'].fill = input_fill; ws['B18'].number_format = '0.0%'

    ws['A19'] = "Terminal Growth Rate"
    ws['B19'] = 0.03
    ws['B19'].font = blue_font; ws['B19'].fill = input_fill; ws['B19'].number_format = '0.0%'

    ws['A20'] = "Discount Factor"
    ws['A21'] = "Present Value of FCF"
    for c in range(3, 8):
        col = get_column_letter(c); year = c - 2
        ws[f'{col}20'] = f"=1/((1+$B$18)^{year})"; ws[f'{col}20'].number_format = '0.00x'
        ws[f'{col}21'] = f"={col}15*{col}20"

    ws['A23'] = "3. ENTERPRISE TO EQUITY VALUATION"
    ws['A23'].fill = subheader_fill; ws['A23'].font = bold_font
    ws['A24'] = "Cumulative PV of FCFs"
    ws['B24'] = "=SUM(C21:G21)"
    ws['A25'] = "Terminal Value"
    ws['B25'] = "=G15*(1+B19)/(B18-B19)"
    ws['A26'] = "PV of Terminal Value"
    ws['B26'] = "=B25*G20"
    ws['A27'] = "Enterprise Value"
    ws['B27'] = "=B24+B26"; ws['B27'].font = bold_font
    ws['A28'] = "Plus: Cash & Equivalents"
    ws['B28'] = total_cash
    ws['A29'] = "Less: Total Debt"
    ws['B29'] = -abs(total_debt)
    ws['A30'] = "Implied Equity Value"
    ws['B30'] = "=B27+B28+B29"; ws['B30'].font = bold_font
    ws['A31'] = "Shares Outstanding"
    ws['B31'] = shares_out

    ws['A33'] = f"IMPLIED SHARE PRICE ({financial_currency})"
    ws['B33'] = "=B30/B31"
    ws['B33'].font = Font(bold=True, size=12); ws['A33'].font = Font(bold=True, size=12)
    ws['B33'].number_format = '"$"#,##0.00' if financial_currency == "USD" else '"₹"#,##0.00'
    ws['B33'].border = top_bottom_border

    if stock_currency != financial_currency:
        ws['A36'] = "4. CROSS-CURRENCY NORMALIZER"
        ws['A36'].fill = subheader_fill; ws['A36'].font = bold_font

        ws['A37'] = f"Live Exchange Rate ({financial_currency} to {stock_currency})"
        ws['B37'] = live_fx_rate
        ws['B37'].font = blue_font 
        ws['B37'].fill = input_fill
        ws['B37'].number_format = '0.00'

        ws['A39'] = f"IMPLIED SHARE PRICE ({stock_currency})"
        ws['B39'] = "=B33*B37"
        ws['B39'].font = Font(bold=True, color="008000", size=14) 
        ws['A39'].font = Font(bold=True, size=12)
        ws['B39'].number_format = '"₹"#,##0.00' if stock_currency == "INR" else '"$"#,##0.00'
        ws['B39'].border = top_bottom_border

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer
# ==========================================
# NATIVE CACHING (YFINANCE CLOUDFLARE COMPATIBLE)
# ==========================================
# Relying purely on Streamlit caching. Custom sessions are removed 
# so yfinance can natively use curl_cffi to bypass Yahoo's firewalls.

@st.cache_data(ttl=3600) 
def fetch_fundamentals(ticker_symbol):
    stock = yf.Ticker(ticker_symbol)
    return stock.info

@st.cache_data(ttl=3600)
def pull_all_data(ticker):
    stock = yf.Ticker(ticker)
    return stock.financials.T, stock.balance_sheet.T, stock.cashflow.T, stock.info

def format_kmb(num):
    if pd.isna(num) or num == 0: return "0"
    is_neg = num < 0
    num = abs(num)
    if num >= 1e9: val = f"{num/1e9:.2f}B"
    elif num >= 1e6: val = f"{num/1e6:.2f}M"
    elif num >= 1e3: val = f"{num/1e3:.2f}K"
    else: val = f"{num:.2f}"
    return f"-{val}" if is_neg else val

# ==========================================
# USER INTERFACE START
# ==========================================
st.title("💹 Financial Research System")
st.markdown("Deep dive into fundamentals and technicals of companies.")

search_col1, search_col2 = st.columns([8, 1])
with search_col1:
    query = st.text_input("Enter any company name (e.g., Microsoft, Infosys, Apple):", label_visibility="collapsed")
with search_col2:
    search_clicked = st.button("Search 🔍", use_container_width=True)


if query:
    ticker = resolve_ticker(query)
    def run_automated_dcf(ticker_obj, current_price):
        try:
            # 1. Pull Raw Accounting Data
            info = ticker_obj.info
            cash_flow = ticker_obj.cashflow
            financials = ticker_obj.financials
            balance_sheet = ticker_obj.balancesheet
            
            # Extract latest year data (safely handling missing data)
            shares_out = info.get('sharesOutstanding', 1)
            total_debt = info.get('totalDebt', 0)
            total_cash = info.get('totalCash', 0)
            
            # Calculate a rough proxy for current year Free Cash Flow
            operating_cash_flow = cash_flow.loc['Operating Cash Flow'].iloc[0]
            capex = cash_flow.loc['Capital Expenditure'].iloc[0]
            current_fcf = operating_cash_flow + capex # CapEx is usually a negative number in accounting
            
            # 2. Automated Assumptions (The AVM Engine)
            projected_growth_rate = 0.08  # Assuming 8% FCF growth for 5 years
            wacc = 0.10                   # 10% Discount Rate
            terminal_growth = 0.03        # 3% Perpetuity growth
            
            # 3. Project 5 Years of FCF
            projected_fcfs = []
            for year in range(1, 6):
                future_fcf = current_fcf * ((1 + projected_growth_rate) ** year)
                projected_fcfs.append(future_fcf)
                
            # 4. Discount FCFs to Present Value
            pv_fcfs = 0
            for year, fcf in enumerate(projected_fcfs, 1):
                pv_fcfs += fcf / ((1 + wacc) ** year)
                
            # 5. Terminal Value
            terminal_value = (projected_fcfs[-1] * (1 + terminal_growth)) / (wacc - terminal_growth)
            pv_terminal_value = terminal_value / ((1 + wacc) ** 5)
            
            # 6. Enterprise Value to Equity Value
            enterprise_value = pv_fcfs + pv_terminal_value
            equity_value = enterprise_value + total_cash - total_debt
            
            implied_share_price = equity_value / shares_out
            
            # Generate DataFrame for Excel Export
            dcf_df = pd.DataFrame({
                "Metric": ["Current FCF", "Year 1 FCF", "Year 2 FCF", "Year 3 FCF", "Year 4 FCF", "Year 5 FCF", "Terminal Value", "Enterprise Value", "Equity Value", "Implied Share Price"],
                "Value (Automated)": [current_fcf, projected_fcfs[0], projected_fcfs[1], projected_fcfs[2], projected_fcfs[3], projected_fcfs[4], terminal_value, enterprise_value, equity_value, implied_share_price]
            })
            
            return implied_share_price, dcf_df
            
        except Exception as e:
            return None, None
    
    # ==========================================
    # PART 1: FUNDAMENTAL ANALYSIS SECTION
    # ==========================================
    with st.spinner(f"Compiling institutional ledgers and running analysis for {ticker}..."):
        try:
            inc, bs, cf, info = pull_all_data(ticker)
            name = info.get('shortName', ticker)
            
            sector, industry = str(info.get('sector', '')), str(info.get('industry', ''))
            is_bank = 'Financial' in sector or 'Bank' in industry or 'Credit' in industry
            
            currency_code = info.get('currency', info.get('financialCurrency', 'USD'))
            if currency_code == 'INR': sym, pdf_sym = '₹', 'Rs. ' 
            elif currency_code == 'EUR': sym, pdf_sym = '€', 'EUR '
            elif currency_code == 'GBP': sym, pdf_sym = '£', 'GBP '
            else: sym, pdf_sym = '$', '$'
            fin_sym = '$' 
            
            price = info.get('currentPrice', info.get('regularMarketPrice', 0))
            
            if HAS_ALTDATA: intrinsic_val = run_altdata_valuation(ticker)
            else: intrinsic_val = info.get('targetMeanPrice', price * 1.08)

            beta_val = info.get('beta', 1.0)
            de_ratio_raw = info.get('debtToEquity', 0)
            de_val = 0.0 if pd.isna(de_ratio_raw) else float(de_ratio_raw)
            
            st.divider()
            
            # --- SAFE MARGIN CALCULATION FOR TOP BAR ---
            try:
                temp_stock = yf.Ticker(ticker)
                top_ni = temp_stock.financials.loc['Net Income'].iloc[0]
                top_rev = temp_stock.financials.loc['Total Revenue'].iloc[0]
                top_margin = (top_ni / top_rev) * 100 if top_rev != 0 else 0
            except:
                top_margin = info.get('profitMargins', 0) * 100

            # --- METRIC DISPLAY ---
            k1, k2, k3, k4, k5, k6 = st.columns(6)

            k1.metric("Market Price", f"{sym}{price:,.2f}")
            k2.metric("Intrinsic Target", f"{sym}{intrinsic_val:,.2f}")
            k3.metric("P/E Ratio", f"{info.get('trailingPE', 'N/A')}")
            k4.metric("Profit Margin", f"{top_margin:.1f}%")

            if is_bank: 
                k5.metric("Debt-to-Equity", "N/A (Bank)")
            elif de_val <= 0.01: 
                k5.metric("Debt-to-Equity", "Debt Free")
            else: 
                k5.metric("Debt-to-Equity", f"{de_val:.2f}")

            k6.metric("Beta (Risk)", f"{beta_val:.2f}")

            st.divider()
            
            # --- THE AI & NLP DASHBOARD SECTION ---
            st.header("Advanced AI & NLP updates")
            nlp_text, consensus_text = "NLP Offline", "Consensus Offline"
            
            ai1, ai2 = st.columns(2)
            with ai1:
                st.markdown("### 📰 Sentiment & NLP Engine")
                if HAS_NLP:
                    nlp_text = str(run_sentiment_analysis(ticker))
                    st.markdown(f"<div class='module-box'>{nlp_text}</div>", unsafe_allow_html=True)
                else: st.warning(f"Error loading dcf_nlp.py: {NLP_ERROR}")
                    
            with ai2:
                st.markdown("### 🕵️‍♂️ Financial Health")
                if HAS_CONSENSUS:
                    consensus_text = str(get_dupont_and_health_metrics(ticker))
                    st.markdown(f"<div class='module-box'>{consensus_text}</div>", unsafe_allow_html=True)
                else: st.warning(f"Error loading dcf_consensus.py: {CONSENSUS_ERROR}")

            st.divider()

            slice_idx = min(4, len(inc))
            try: years = [str(date.year) for date in inc.index[:slice_idx]][::-1]
            except: years = [f"Year {i}" for i in range(1, slice_idx+1)]

            rev = inc.get('Total Revenue', inc.get('Operating Revenue', inc.iloc[:, 0])).iloc[:slice_idx][::-1].fillna(0)
            ni = inc.get('Net Income', rev * 0.1).iloc[:slice_idx][::-1].fillna(0)
            gp = inc.get('Gross Profit', rev * 0.4).iloc[:slice_idx][::-1].fillna(0)
            op_inc = inc.get('Operating Income', rev * 0.2).iloc[:slice_idx][::-1].fillna(0)
            ebitda = inc.get('EBITDA', op_inc * 1.2).iloc[:slice_idx][::-1].fillna(0)
            interest = inc.get('Interest Expense', pd.Series([-1]*len(inc))).iloc[:slice_idx][::-1].fillna(-1).abs().replace(0, 1)
            
            debt = bs.get('Total Debt', pd.Series([0]*len(bs))).iloc[:slice_idx][::-1].fillna(0)
            equity = bs.get('Stockholders Equity', pd.Series([1]*len(bs))).iloc[:slice_idx][::-1].fillna(1)
            assets = bs.get('Total Assets', pd.Series([1]*len(bs))).iloc[:slice_idx][::-1].fillna(1)
            liab = bs.get('Total Liabilities Net Minority Interest', bs.get('Total Liabilities', pd.Series([1]*len(bs)))).iloc[:slice_idx][::-1].fillna(1)

            ca_row, cl_row = bs.get('Current Assets'), bs.get('Current Liabilities')
            if ca_row is None: ca_row = bs.get('Total Current Assets')
            if cl_row is None: cl_row = bs.get('Total Current Liabilities')
            has_wc_data = ca_row is not None and cl_row is not None and not ca_row.isna().all()
            
            if has_wc_data:
                curr_assets = ca_row.iloc[:slice_idx][::-1].fillna(0)
                curr_liab = cl_row.iloc[:slice_idx][::-1].fillna(0)
            else:
                curr_assets, curr_liab = pd.Series([0]*slice_idx, index=years), pd.Series([0]*slice_idx, index=years)

            ocf = cf.get('Operating Cash Flow', ni * 1.1).iloc[:slice_idx][::-1].fillna(0)
            capex = cf.get('Capital Expenditure', rev * -0.05).iloc[:slice_idx][::-1].fillna(0).abs()
            fcf = cf.get('Free Cash Flow', ocf - capex).iloc[:slice_idx][::-1].fillna(0)

            c1, c2 = st.columns(2)
            with c1:
                st.header("1. Top & Bottom Line Growth")
                fig1 = go.Figure()
                fig1.add_trace(go.Bar(x=years, y=rev/1e9, name="Revenue", marker_color='#3b82f6'))
                fig1.add_trace(go.Scatter(x=years, y=ni/1e9, name="Net Income", line=dict(color='#10b981', width=4)))
                fig1.update_layout(template="plotly_dark", height=350, margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title=f"Amount ({fin_sym} Billions)")
                st.plotly_chart(fig1, use_container_width=True)
                st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> Tracks total sales (Blue) against actual money kept after all expenses (Green).<br><br><b>Trend:</b> Revenue reached {fin_sym}{rev.iloc[-1]/1e9:.2f}B in the most recent year.<br><br><b>Conclusion:</b> By comparing the trajectory, we see if {name} is sacrificing profit to gain market share. A rising green line means management is scaling operations profitably without bleeding cash.</div>", unsafe_allow_html=True)

            with c2:
                st.header("2. Margin Efficiency Trajectory")
                fig2 = go.Figure()
                fig2.add_trace(go.Scatter(x=years, y=(gp/rev)*100, name="Gross Margin %", line=dict(color='#f59e0b', width=3)))
                fig2.add_trace(go.Scatter(x=years, y=(op_inc/rev)*100, name="Operating Margin %", line=dict(color='#8b5cf6', width=3)))
                fig2.add_trace(go.Scatter(x=years, y=(ni/rev)*100, name="Net Margin %", fill='tozeroy', line_color='#10b981'))
                fig2.update_layout(template="plotly_dark", height=350, margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title="Percentage (%)")
                st.plotly_chart(fig2, use_container_width=True)
                st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> Margins measure operational efficiency. Gross is after manufacturing, Operating is after overhead, Net is after taxes.<br><br><b>Trend:</b> The final net margin shifted to {(ni.iloc[-1]/rev.iloc[-1])*100:.1f}% this year.<br><br><b>Conclusion:</b> The gap between Gross (Orange) and Operating (Purple) is what they spend on administration. A stable green area proves strong pricing power.</div>", unsafe_allow_html=True)

            c3, c4 = st.columns(2)
            with c3:
                st.header("3. Quality of Earnings")
                fig3 = go.Figure()
                fig3.add_trace(go.Bar(x=years, y=ni/1e9, name="Net Income", marker_color='#64748b'))
                fig3.add_trace(go.Bar(x=years, y=ocf/1e9, name="Operating Cash Flow", marker_color='#14b8a6'))
                fig3.update_layout(template="plotly_dark", height=350, barmode='group', margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title=f"Amount ({fin_sym} Billions)")
                st.plotly_chart(fig3, use_container_width=True)
                st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> Net Income is easily manipulated by accounting rules. Operating Cash Flow is the physical money hitting the bank.<br><br><b>Trend:</b> Cash Flow sits at {fin_sym}{ocf.iloc[-1]/1e9:.2f}B vs Net Income of {fin_sym}{ni.iloc[-1]/1e9:.2f}B.<br><br><b>Conclusion:</b> Because cash flow is {'higher' if ocf.iloc[-1]>ni.iloc[-1] else 'lower'} than paper profits, the earnings are backed by true liquidity. High-quality businesses routinely show higher operating cash than net income.</div>", unsafe_allow_html=True)

            with c4:
                st.header("4. CapEx & Free Cash Flow")
                fig4 = go.Figure()
                fig4.add_trace(go.Bar(x=years, y=ocf/1e9, name="Cash Generated", marker_color='#3b82f6'))
                fig4.add_trace(go.Bar(x=years, y=-capex/1e9, name="CapEx", marker_color='#ef4444'))
                fig4.add_trace(go.Scatter(x=years, y=fcf/1e9, name="Free Cash Flow", line=dict(color='#10b981', width=4)))
                fig4.update_layout(template="plotly_dark", height=350, barmode='relative', margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title=f"Amount ({fin_sym} Billions)")
                st.plotly_chart(fig4, use_container_width=True)
                st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> CapEx (Red) is money reinvested into physical assets. Free Cash Flow (Green) is what remains for shareholders.<br><br><b>Trend:</b> Reinvestment required {fin_sym}{capex.iloc[-1]/1e9:.2f}B this year.<br><br><b>Conclusion:</b> The green line represents the true value of the business. Generating {fin_sym}{fcf.iloc[-1]/1e9:.2f}B in pure free cash gives management massive power to buy back stock, acquire competitors, or pay dividends.</div>", unsafe_allow_html=True)

            c5, c6 = st.columns(2)
            with c5:
                st.header("5. Short-Term Liquidity")
                if is_bank or not has_wc_data:
                    fig5 = go.Figure(go.Scatter(x=years, y=[0]*len(years), mode='lines', line=dict(color='#374151')))
                    fig5.update_layout(template="plotly_dark", height=350, margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title="Ratio")
                    st.plotly_chart(fig5, use_container_width=True)
                    st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> The Current Ratio measures if the company has enough liquid assets to pay debts due within 12 months.<br><br><b>Trend:</b> Data Unavailable.<br><br><b>Conclusion:</b> Current/Non-Current asset categorization is structurally inapplicable or unavailable for this ticker in standard financial feeds.</div>", unsafe_allow_html=True)
                else:
                    cr = np.where(curr_liab == 0, 0, curr_assets / curr_liab)
                    fig5 = go.Figure(go.Scatter(x=years, y=cr, mode='lines+markers', line=dict(color='#facc15', width=4)))
                    fig5.update_layout(template="plotly_dark", height=350, margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title="Ratio")
                    st.plotly_chart(fig5, use_container_width=True)
                    st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> The Current Ratio measures if the company has enough liquid assets to pay debts due within 12 months.<br><br><b>Trend:</b> The ratio stabilized at {cr[-1]:.2f}x.<br><br><b>Conclusion:</b> A ratio above 1.0 indicates adequate liquidity to survive the year without external bridging.</div>", unsafe_allow_html=True)

            with c6:
                st.header("6. Long-Term Solvency")
                fig6 = go.Figure()
                fig6.add_trace(go.Bar(x=years, y=equity/1e9, name="Equity", marker_color='#10b981'))
                fig6.add_trace(go.Bar(x=years, y=debt/1e9, name="Debt", marker_color='#ef4444'))
                fig6.update_layout(template="plotly_dark", height=350, barmode='group', margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title=f"Amount ({fin_sym} Billions)")
                st.plotly_chart(fig6, use_container_width=True)
                if is_bank:
                    st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> Shareholder Equity vs. Corporate Debt.<br><br><b>Trend:</b> N/A Corporate Debt for Banks.<br><br><b>Conclusion:</b> Financial institutions utilize customer deposits as their primary leverage rather than standard corporate debt. Therefore, debt bars are structurally zero or minimal.</div>", unsafe_allow_html=True)
                else:
                    st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> A direct visual comparison of shareholder money (Green) vs. borrowed bank money (Red).<br><br><b>Trend:</b> Total corporate debt is {fin_sym}{debt.iloc[-1]/1e9:.2f}B.<br><br><b>Conclusion:</b> The Debt-to-Equity ratio reveals the firm's leverage profile. Less reliance on red bars means the company is self-funded and highly resilient to macroeconomic shocks.</div>", unsafe_allow_html=True)

            c7, c8 = st.columns(2)
            with c7:
                st.header("7. EBITDA Trajectory")
                fig7 = go.Figure(go.Bar(x=years, y=ebitda/1e9, marker_color='#8b5cf6'))
                fig7.update_layout(template="plotly_dark", height=350, margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title=f"Amount ({fin_sym} Billions)")
                st.plotly_chart(fig7, use_container_width=True)
                st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> EBITDA stands for Earnings Before Interest, Taxes, Depreciation, and Amortization.<br><br><b>Trend:</b> Core operational profit reached {fin_sym}{ebitda.iloc[-1]/1e9:.2f}B.<br><br><b>Conclusion:</b> This strips away accounting depreciation and taxes to show the raw, pure engine of the business before external factors.</div>", unsafe_allow_html=True)

            with c8:
                st.header("8. Debt Servicing (Interest Coverage)")
                icr = op_inc / interest
                fig8 = go.Figure(go.Scatter(x=years, y=icr, fill='tozeroy', line_color='#ec4899'))
                fig8.add_hline(y=1.5, line_dash="dash", line_color="red")
                fig8.update_layout(template="plotly_dark", height=350, margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title="Ratio")
                st.plotly_chart(fig8, use_container_width=True)
                st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> Shows how many times over the company can pay the interest on its debt using its operating profit.<br><br><b>Trend:</b> The ratio is currently {icr.iloc[-1]:.1f}x.<br><br><b>Conclusion:</b> {name} generates {icr.iloc[-1]:.1f}x more operating profit than it needs to pay its bankers. Lenders view anything above the red line as safe.</div>", unsafe_allow_html=True)

            c9, c10 = st.columns(2)
            with c9:
                st.header("9. Return on Assets & Equity")
                fig9 = go.Figure()
                fig9.add_trace(go.Scatter(x=years, y=(ni/assets)*100, name="ROA %", line=dict(color='#06b6d4', width=3)))
                fig9.add_trace(go.Scatter(x=years, y=(ni/equity)*100, name="ROE %", line=dict(color='#3b82f6', width=3)))
                fig9.update_layout(template="plotly_dark", height=350, margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title="Percentage (%)")
                st.plotly_chart(fig9, use_container_width=True)
                st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> ROA (Teal) shows profit generated from total assets. ROE (Blue) shows profit generated specifically from shareholder money.<br><br><b>Trend:</b> ROE stands at {(ni.iloc[-1]/equity.iloc[-1])*100:.1f}%.<br><br><b>Conclusion:</b> This proves how effectively management is compounding investor capital. A rising blue line is the ultimate goal for shareholders.</div>", unsafe_allow_html=True)

            with c10:
                st.header("10. Net Working Capital")
                if is_bank or not has_wc_data:
                    fig10 = go.Figure(go.Bar(x=years, y=[0]*len(years), marker_color='#374151'))
                    fig10.update_layout(template="plotly_dark", height=350, margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title=f"Amount ({fin_sym} Billions)")
                    st.plotly_chart(fig10, use_container_width=True)
                    st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> Current Assets minus Current Liabilities.<br><br><b>Trend:</b> Data Unavailable.<br><br><b>Conclusion:</b> Metrics are structurally inapplicable or explicitly missing for this specific ticker.</div>", unsafe_allow_html=True)
                else:
                    nwc = curr_assets - curr_liab
                    fig10 = go.Figure(go.Bar(x=years, y=nwc/1e9, marker_color=np.where(nwc>0, '#10b981', '#ef4444')))
                    fig10.update_layout(template="plotly_dark", height=350, margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title=f"Amount ({fin_sym} Billions)")
                    st.plotly_chart(fig10, use_container_width=True)
                    st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> Current Assets minus Current Liabilities. This measures the short-term liquidity runway.<br><br><b>Trend:</b> Working Capital sits at {fin_sym}{nwc.iloc[-1]/1e9:.2f}B.<br><br><b>Conclusion:</b> A positive (Green) number gives the company massive flexibility to fund daily operations without taking loans. Red indicates a potential cash crunch.</div>", unsafe_allow_html=True)

            c11, c12 = st.columns(2)
            with c11:
                st.header("11. Cost Structure (The Revenue Dollar)")
                cogs = rev.iloc[-1] - gp.iloc[-1]
                opex = gp.iloc[-1] - op_inc.iloc[-1]
                taxes_other = op_inc.iloc[-1] - ni.iloc[-1]
                
                raw_values = [max(0, cogs), max(0, opex), max(0, taxes_other), max(0, ni.iloc[-1])]
                labels = ['Cost of Goods', 'Operating Expenses', 'Taxes/Other', 'Net Profit Retained']
                formatted_text = [format_kmb(v) for v in raw_values]

                fig11 = go.Figure(go.Pie(
                    labels=labels, values=raw_values, hole=0.4,
                    marker_colors=['#f59e0b', '#8b5cf6', '#ef4444', '#10b981'],
                    textinfo='percent', hovertext=formatted_text, hoverinfo='label+text+percent' 
                ))
                fig11.update_layout(
                    template="plotly_dark", height=350, margin=dict(b=0, t=30),
                    legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5)
                )
                st.plotly_chart(fig11, use_container_width=True)
                st.markdown(f"<div class='conclusion-box'><b>Concept:</b> Breaks down exactly where every dollar of {years[-1]} revenue went.<br><br><b>Trend:</b> The green slice represents the final {(ni.iloc[-1]/rev.iloc[-1])*100:.1f}% kept as profit.<br><br><b>Conclusion:</b> This visualizes the company's cost burden. A larger green slice relative to orange and purple indicates a highly efficient business model.</div>", unsafe_allow_html=True)
            with c12:
                st.header("12. Asset vs. Liability Composition")
                fig12 = go.Figure()
                fig12.add_trace(go.Bar(x=years, y=assets/1e9, name="Total Assets", marker_color='#3b82f6'))
                fig12.add_trace(go.Bar(x=years, y=liab/1e9, name="Total Liabilities", marker_color='#ef4444'))
                fig12.update_layout(template="plotly_dark", height=350, barmode='group', margin=dict(b=0), xaxis_title="Fiscal Year", yaxis_title=f"Amount ({fin_sym} Billions)")
                st.plotly_chart(fig12, use_container_width=True)
                st.markdown(f"<div class='conclusion-box'><b>Agenda:</b> Compares everything the company owns (Assets) against everything it owes (Liabilities).<br><br><b>Trend:</b> Total Assets reached {fin_sym}{assets.iloc[-1]/1e9:.2f}B.<br><br><b>Conclusion:</b> As long as the Blue bars grow faster than the Red bars, the company is building intrinsic equity value and increasing its net worth.</div>", unsafe_allow_html=True)

            st.markdown("<div class='custom-line'></div>", unsafe_allow_html=True)
            st.header("13. Traditional M&A & Capital Velocity")

            acquisitions = cf.get('Net Business Purchase And Sale', pd.Series([0]*len(cf))).iloc[:slice_idx][::-1].fillna(0).abs()
            total_investment = cf.get('Investing Cash Flow', rev * -0.1).iloc[:slice_idx][::-1].fillna(0).abs()

            m1, m2 = st.columns([1, 2])

            with m1:
                st.markdown("### M&A Strategic Summary")
                avg_acquisition_spend = acquisitions.mean() / 1e9
                intensity_ratio = (acquisitions.sum() / total_investment.sum()) * 100 if total_investment.sum() != 0 else 0
                
                st.write(f"""
                This section monitors the **Inorganic Growth Engine**. By analyzing traditional M&A data, we track how aggressively {name} is buying market share.
                
                - **Avg. Annual Acquisition Spend:** {sym}{avg_acquisition_spend:.2f}B
                - **Capital Reinvestment Intensity:** {intensity_ratio:.1f}%
                """)
                
                if intensity_ratio > 30:
                    st.info("💡 **Strategy Note:** High M&A Intensity. The company is in an 'Aggressive Consolidator' phase.")
                else:
                    st.info("💡 **Strategy Note:** Low M&A Intensity. Growth is currently driven by internal R&D and organic operations.")

            with m2:
                fig13 = go.Figure()
                fig13.add_trace(go.Scatter(x=years, y=total_investment/1e9, name="Total Investing Outlay", fill='tonexty', line_color='#6366f1'))
                fig13.add_trace(go.Bar(x=years, y=acquisitions/1e9, name="M&A Spend", marker_color='#f59e0b'))
                
                fig13.update_layout(
                    template="plotly_dark", height=400, margin=dict(b=0),
                    yaxis_title=f"Amount in {sym} Billions",
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                st.plotly_chart(fig13, use_container_width=True)

            st.markdown(f"<div class='conclusion-box'><b>Concept:</b> Compares total capital invested (Purple) against specific M&A acquisitions (Orange).<br><br><b>Trend:</b> M&A activity reached {sym}{acquisitions.iloc[-1]/1e9:.2f}B this period.<br><br><b>Conclusion:</b> Monitoring the gap between these two figures reveals if the company is growing through synergy (buying others) or execution (building internally).</div>", unsafe_allow_html=True)
            
            # --- MASSIVE AUTOMATED PDF REPORT & EXCEL EXPORT ---
            c_exp1, c_exp2 = st.columns(2)
        
            def generate_deep_pdf():
                sentiment_tag = "NEUTRAL" 
                if "BULLISH" in nlp_text: sentiment_tag = "BULLISH"
                elif "BEARISH" in nlp_text: sentiment_tag = "BEARISH"

                roe_val = info.get('returnOnEquity', 0) * 100
                try:
                    raw_equity = bs['Stockholders Equity'].iloc[-1]
                    safe_equity = 0 if pd.isna(raw_equity) else (raw_equity / 1e9)
                except:
                    safe_equity = 0

                equity_val = bs.get('Stockholders Equity', pd.Series([0])).iloc[-1]
                if pd.isna(equity_val): equity_val = 0
                
                is_positive = ni.iloc[-1] > 0
                fcf_is_positive = fcf.iloc[-1] > 0

                health_status = "robust operational integrity" if is_positive else "significant operational challenges"
                sustainability_msg = "The gap between cash made and what is needed shows great long-term health." if is_positive else "The gap between cash coming in and rising bills shows serious risks for the business."
                management_msg = "management is putting cash into projects that grow the business." if is_positive else "management is struggling to keep the business profitable against high costs."
                final_action = "buy other companies or reward shareholders with extra dividends." if is_positive else "focus on cutting debt and saving cash to keep the business running."

                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", 'B', 16)
                pdf.multi_cell(0, 10, f"FINANCIAL RESEARCH BREAKDOWN: {name.upper()}", align='C')
                
                pdf.set_font("Arial", 'I', 10)
                pdf.cell(0, 10, f"Market Price: {pdf_sym}{price:,.2f} | Consensus Target: {pdf_sym}{intrinsic_val:,.2f} | Base Currency: {currency_code}", ln=True, align='C')
                
                pdf.line(10, 32, 200, 32)
                pdf.ln(12)
                
                if isinstance(nlp_text, tuple): safe_nlp = clean_for_pdf(nlp_text[0])
                else: safe_nlp = clean_for_pdf(nlp_text)
                safe_consensus = clean_for_pdf(consensus_text)
                
                sentiment_interpretation = f"What does the sentiment mean? The NLP engine found a {sentiment_tag} mood in the news. This means the latest stories are highlighting growth chances or risks that aren't fully shown in the stock price yet. This helps investors see if the stock has a safety net or if it might keep falling."
                financial_interpretation = f"What does the financial health mean? With a ROE of {roe_val:.2f}%, {management_msg} For anyone reading this, it shows whether the company is actually building value or just using up its resources to stay afloat."
                
                sections = {
                    "1. Executive Summary & Intrinsic Valuation": f"This document provides a comprehensive financial breakdown of {name}. The asset is currently trading at a market price of {pdf_sym}{price:,.2f}. Based on alternative data and Wall Street consensus, the intrinsic target price is modeled at {pdf_sym}{intrinsic_val:,.2f}. Our analysis strips away standard accounting noise to focus on operational cash flow, capital structure, and true margin efficiency to determine the long-term viability of the underlying business model.",
                    "2. AI Sentiment & NLP Insights": f"{safe_nlp}",
                    "3. Consensus Forensic Breakdown": f"{safe_consensus}",
                    "4. Top-Line Trajectory & Margin Efficiency": f"Growth metrics over the observed {slice_idx}-year period reveal that total revenue successfully scaled to {pdf_sym}{rev.iloc[-1]/1e9:.2f} Billion. This top-line growth is supported by a gross margin of {(gp.iloc[-1]/rev.iloc[-1])*100:.1f}%. By managing their office and daily costs, the firm secured an operating margin of {(op_inc.iloc[-1]/rev.iloc[-1])*100:.1f}%. After all taxes, the final profit margin left for shareholders is {(ni.iloc[-1]/rev.iloc[-1])*100:.1f}%.",
                    "5. Earnings Quality Forensics": f"Real cash (Operating Cash Flow) is harder to fake than paper profits. This analysis shows {name} produced {pdf_sym}{ocf.iloc[-1]/1e9:.2f} Billion in actual cash. Since this cash amount is {'higher' if ocf.iloc[-1]>ni.iloc[-1] else 'lower'} than the reported profit of {pdf_sym}{ni.iloc[-1]/1e9:.2f} Billion, we can see how honest and strong their accounting really is.",
                    "6. Capital Reinvestment (CapEx) & Free Cash Flow": f"To keep the business running and growing, the company spent {pdf_sym}{capex.iloc[-1]/1e9:.2f} Billion on equipment and upgrades. After paying for these, the leftover 'Free Cash Flow' is {pdf_sym}{fcf.iloc[-1]/1e9:.2f} Billion. This {'helps' if fcf_is_positive else 'hurts'} the company's ability to stay independent and grow without needing to borrow more money.",
                    "7. Capital Structure, Liquidity & Default Risk": f"The balance sheet shows the company has {pdf_sym}{safe_equity:.2f} Billion in value for owners against debts of {pdf_sym}{info.get('totalDebt', 0)/1e9:.2f} Billion. This tells us if the company has enough money in the bank to pay its bills over the next year without running into trouble.",
                    "8. M&A Strategic Velocity & Portfolio Impact": f"The company spends about {pdf_sym}{acquisitions.mean()/1e9:.2f}B a year buying other businesses. This M&A spending makes up {(acquisitions.sum()/total_investment.sum())*100:.1f}% of their total investment. This shows if they are growing by being better at their job or just by buying out competitors.",
                    "9. Final Strategic Verdict": f"The multi-year data confirms that {name} shows {health_status}. {sustainability_msg} \n\n{sentiment_interpretation} \n\n{financial_interpretation} \n\nIn light of these metrics, management should {final_action} as a priority to secure the company's long-term future."
                }
                
                for title, text in sections.items():
                    pdf.set_font("Arial", 'B', 12)
                    pdf.cell(0, 10, title, ln=True)
                    pdf.set_font("Arial", '', 11)
                    pdf.multi_cell(0, 6, text.encode('latin-1', 'replace').decode('latin-1'))
                    pdf.ln(8)
                    
                return pdf.output(dest="S").encode("latin-1")

            def generate_excel():
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df_formatted = pd.DataFrame({
                        'Revenue': rev.apply(lambda x: fin_sym + format_kmb(x).replace('$', '')),
                        'Net Income': ni.apply(lambda x: fin_sym + format_kmb(x).replace('$', '')),
                        'Operating Cash Flow': ocf.apply(lambda x: fin_sym + format_kmb(x).replace('$', '')),
                        'CapEx': capex.apply(lambda x: fin_sym + format_kmb(x).replace('$', '')),
                        'Free Cash Flow (UFCF)': fcf.apply(lambda x: fin_sym + format_kmb(x).replace('$', '')),
                        'Total Equity': equity.apply(lambda x: fin_sym + format_kmb(x).replace('$', '')),
                        'Total Debt': debt.apply(lambda x: fin_sym + format_kmb(x).replace('$', ''))
                    })
                    
                    df_raw = pd.DataFrame({
                        'Revenue': rev/1e9, 
                        'Gross Profit': gp/1e9,
                        'Net Income': ni/1e9,
                        'Op Cash Flow': ocf/1e9,
                        'CapEx': capex/1e9,
                        'Free Cash Flow': fcf/1e9,
                        'Total Equity': equity/1e9,
                        'Total Debt': debt/1e9
                    })
                    
                    df_formatted.to_excel(writer, sheet_name='Financial Model')
                    df_raw.to_excel(writer, sheet_name='Chart Data')
                    
                    workbook = writer.book
                    worksheet = writer.sheets['Financial Model']
                    max_row = len(years) + 1

                    chart1 = workbook.add_chart({'type': 'column'})
                    chart1.add_series({'name': "='Chart Data'!$B$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$B$2:$B${max_row}", 'fill': {'color': '#3b82f6'}})
                    chart1.add_series({'name': "='Chart Data'!$D$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$D$2:$D${max_row}", 'fill': {'color': '#10b981'}})
                    chart1.set_title({'name': f'Revenue vs Net Income ({fin_sym} Billions)'})
                    worksheet.insert_chart('J2', chart1)

                    chart2 = workbook.add_chart({'type': 'column'})
                    chart2.add_series({'name': "='Chart Data'!$E$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$E$2:$E${max_row}", 'fill': {'color': '#14b8a6'}})
                    chart2.add_series({'name': "='Chart Data'!$F$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$F$2:$F${max_row}", 'fill': {'color': '#ef4444'}})
                    chart2.add_series({'name': "='Chart Data'!$G$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$G$2:$G${max_row}", 'type': 'line', 'line': {'color': '#10b981', 'width': 3}})
                    chart2.set_title({'name': f'Cash Flow Engine ({fin_sym} Billions)'})
                    worksheet.insert_chart('J17', chart2)

                    chart3 = workbook.add_chart({'type': 'column'})
                    chart3.add_series({'name': "='Chart Data'!$H$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$H$2:$H${max_row}", 'fill': {'color': '#10b981'}})
                    chart3.add_series({'name': "='Chart Data'!$I$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$I$2:$I${max_row}", 'fill': {'color': '#ef4444'}})
                    chart3.set_title({'name': f'Capital Structure ({fin_sym} Billions)'})
                    worksheet.insert_chart('R2', chart3)

                    chart4 = workbook.add_chart({'type': 'column'})
                    chart4.add_series({'name': "='Chart Data'!$B$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$B$2:$B${max_row}", 'fill': {'color': '#3b82f6'}})
                    chart4.add_series({'name': "='Chart Data'!$C$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$C$2:$C${max_row}", 'fill': {'color': '#f59e0b'}})
                    chart4.set_title({'name': f'Revenue vs Gross Profit ({fin_sym} Billions)'})
                    worksheet.insert_chart('R17', chart4)
                    
                    df_def = pd.DataFrame({
                        'Metric': ['Revenue', 'CapEx', 'Free Cash Flow (UFCF)', 'Current Ratio', 'Debt-to-Equity'],
                        'Analyst Definition': [
                            'Total top-line sales generated by the core business.',
                            'Capital Expenditures: Money spent on buying or upgrading physical assets.',
                            'Unlevered Free Cash Flow: Pure cash left over after expenses and CapEx.',
                            'Short-term liquidity. Current Assets divided by Current Liabilities.',
                            'Total Debt divided by Shareholder Equity. Measures long-term risk.'
                        ]
                    })
                    df_def.to_excel(writer, sheet_name='Metric Guide', index=False)
                    
                return output.getvalue()
            
            
        except Exception as e:
            st.error(f"Critical Error Mining Fundamental Data: {e}")
        
        # import io
        # import openpyxl
        # from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        # from openpyxl.utils import get_column_letter

        # def generate_dcf_excel(ticker_symbol, base_revenue, shares_out, total_cash, total_debt):
        #     wb = openpyxl.Workbook()
        #     ws = wb.active
        #     ws.title = "Proprietary DCF Model"

        #     # --- Setup Styles ---
        #     header_fill = PatternFill(start_color="1B2631", end_color="1B2631", fill_type="solid")
        #     subheader_fill = PatternFill(start_color="EAEDED", end_color="EAEDED", fill_type="solid")
        #     input_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
            
        #     white_font = Font(color="FFFFFF", bold=True)
        #     blue_font = Font(color="0000FF") # Hardcoded inputs
        #     bold_font = Font(bold=True)
            
        #     top_bottom_border = Border(top=Side(style='thin'), bottom=Side(style='double'))

        #     ws.column_dimensions['A'].width = 38
        #     for i in range(2, 9):
        #         ws.column_dimensions[get_column_letter(i)].width = 16

        #     # --- Headers ---
        #     ws['A1'] = f"DISCOUNTED CASH FLOW (DCF) MODEL - {ticker_symbol.upper()}"
        #     ws['A1'].font = Font(bold=True, size=14)
        #     ws['A2'] = "Format: Institutional Standard (Unlevered FCF)"
            
        #     headers = ["(Figures in Millions)", "Year 0 (Current)", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
        #     for col_num, header in enumerate(headers, 1):
        #         cell = ws.cell(row=4, column=col_num, value=header)
        #         cell.fill = header_fill
        #         cell.font = white_font

        #     # --- 1. UNLEVERED FCF BUILD ---
        #     ws['A5'] = "1. UNLEVERED FREE CASH FLOW BUILD"
        #     ws['A5'].fill = subheader_fill; ws['A5'].font = bold_font

        #     # Revenue
        #     ws['A6'] = "Total Revenue"
        #     ws['B6'] = base_revenue
        #     ws['A7'] = "  % YoY Growth"
            
        #     for c in range(3, 8):
        #         col = get_column_letter(c); prev_col = get_column_letter(c-1)
        #         ws[f'{col}7'] = 0.08  # 8% Growth Input
        #         ws[f'{col}7'].font = blue_font; ws[f'{col}7'].fill = input_fill; ws[f'{col}7'].number_format = '0.0%'
        #         ws[f'{col}6'] = f"={prev_col}6*(1+{col}7)"
        #         ws[f'{col}6'].number_format = '#,##0'

        #     # EBIT & Taxes
        #     ws['A8'] = "EBIT (Operating Income)"
        #     ws['A9'] = "  % EBIT Margin"
        #     ws['A10'] = "Less: Taxes (25%)"
        #     ws['A11'] = "EBIAT (Net Operating Profit After Tax)"
            
        #     for c in range(2, 8):
        #         col = get_column_letter(c)
        #         ws[f'{col}9'] = 0.20 # 20% Margin Input
        #         ws[f'{col}9'].font = blue_font; ws[f'{col}9'].fill = input_fill; ws[f'{col}9'].number_format = '0.0%'
        #         ws[f'{col}8'] = f"={col}6*{col}9"
        #         ws[f'{col}10'] = f"=-{col}8*0.25"
        #         ws[f'{col}11'] = f"={col}8+{col}10"
        #         ws[f'{col}11'].font = bold_font

        #     # D&A, CapEx, NWC
        #     ws['A12'] = "Plus: Depreciation & Amortization"
        #     ws['A13'] = "Less: Capital Expenditure"
        #     ws['A14'] = "Less: Change in Net Working Capital"
        #     ws['A15'] = "UNLEVERED FREE CASH FLOW"
            
        #     for c in range(2, 8):
        #         col = get_column_letter(c)
        #         ws[f'{col}12'] = f"={col}6*0.04" # D&A assumption
        #         ws[f'{col}13'] = f"=-{col}6*0.05" # CapEx assumption
        #         ws[f'{col}14'] = f"=-{col}6*0.01" # NWC assumption
        #         ws[f'{col}15'] = f"=SUM({col}11:{col}14)"
        #         ws[f'{col}15'].font = bold_font
        #         ws[f'{col}15'].border = top_bottom_border

        #     # --- 2. WACC & DISCOUNTING ---
        #     ws['A17'] = "2. DISCOUNTING & TERMINAL VALUE"
        #     ws['A17'].fill = subheader_fill; ws['A17'].font = bold_font
            
        #     ws['A18'] = "Weighted Average Cost of Capital (WACC)"
        #     ws['B18'] = 0.10
        #     ws['B18'].font = blue_font; ws['B18'].fill = input_fill; ws['B18'].number_format = '0.0%'
            
        #     ws['A19'] = "Terminal Growth Rate"
        #     ws['B19'] = 0.03
        #     ws['B19'].font = blue_font; ws['B19'].fill = input_fill; ws['B19'].number_format = '0.0%'

        #     ws['A20'] = "Discount Factor"
        #     ws['A21'] = "Present Value of FCF"
        #     for c in range(3, 8):
        #         col = get_column_letter(c); year = c - 2
        #         ws[f'{col}20'] = f"=1/((1+$B$18)^{year})"
        #         ws[f'{col}20'].number_format = '0.00x'
        #         ws[f'{col}21'] = f"={col}15*{col}20"

        #     # --- 3. VALUATION ---
        #     ws['A23'] = "3. ENTERPRISE TO EQUITY VALUATION"
        #     ws['A23'].fill = subheader_fill; ws['A23'].font = bold_font

        #     ws['A24'] = "Cumulative PV of FCFs"
        #     ws['B24'] = "=SUM(C21:G21)"
            
        #     ws['A25'] = "Terminal Value"
        #     ws['B25'] = "=G15*(1+B19)/(B18-B19)"
            
        #     ws['A26'] = "PV of Terminal Value"
        #     ws['B26'] = "=B25*G20"
            
        #     ws['A27'] = "Enterprise Value"
        #     ws['B27'] = "=B24+B26"
        #     ws['B27'].font = bold_font

        #     ws['A28'] = "Plus: Cash & Equivalents"
        #     ws['B28'] = total_cash
        #     ws['A29'] = "Less: Total Debt"
        #     ws['B29'] = -abs(total_debt) # Ensures it subtracts properly
            
        #     ws['A30'] = "Implied Equity Value"
        #     ws['B30'] = "=B27+B28+B29"
        #     ws['B30'].font = bold_font

        #     ws['A31'] = "Shares Outstanding"
        #     ws['B31'] = shares_out
            
        #     ws['A33'] = "IMPLIED SHARE PRICE"
        #     ws['B33'] = "=B30/B31"
        #     ws['B33'].font = Font(bold=True, size=12)
        #     ws['A33'].font = Font(bold=True, size=12)
        #     ws['B33'].number_format = '"₹"#,##0.00'
        #     ws['B33'].border = top_bottom_border

        #     # Safely extract financial variables EXCLUSIVELY from the standardized 'info' dict
        # # This prevents unit mismatch errors between Revenue, Debt, and Shares
        #     current_revenue = info.get('totalRevenue', 1000000000) 
        #     shares_out = info.get('sharesOutstanding', 1)
        #     total_cash = info.get('totalCash', 0)
        #     total_debt = info.get('totalDebt', 0)

        #     excel_buffer = io.BytesIO()
        #     wb.save(excel_buffer)
        #     excel_buffer.seek(0)
        #     return excel_buffer
# ==========================================
        # PART 2: THE CUSTOM PRO TERMINAL (TIMEZONE & VOLUME FIXED)
        # ==========================================
        st.markdown("---") 
        st.markdown("### 📈 Market Eye")
        st.write("Decoding global liquidity through time synced volume.")

        # 1. Trading Intervals
        interval_options = {
            "1 Minute (Scalping)": ("1m", "5d"),
            "5 Minutes": ("5m", "1mo"),
            "15 Minutes": ("15m", "1mo"),
            "30 Minutes": ("30m", "1mo"),
            "1 Hour": ("1h", "1mo")
        }

        selected_interval = st.selectbox("Select Trading Interval:", list(interval_options.keys()), index=1)
        yf_interval, yf_period = interval_options[selected_interval]

        with st.spinner(f"Rendering pro-grade terminal for {ticker}..."):
            try:
                # 2. Fetch Free Intraday Data
                df_tech = yf.download(ticker, period=yf_period, interval=yf_interval, progress=False)

                if isinstance(df_tech.columns, pd.MultiIndex):
                    df_tech.columns = df_tech.columns.get_level_values(0)

                if not df_tech.empty and len(df_tech) > 50:
                    
                    # ==========================================
                    # THE TIMEZONE FIX (IST vs EST)
                    # ==========================================
                    # Make sure the data has a timezone, then convert it based on the exchange
                    if df_tech.index.tz is None:
                        df_tech.index = df_tech.index.tz_localize('UTC')
                        
                    if ticker.endswith('.NS') or ticker.endswith('.BO'):
                        df_tech.index = df_tech.index.tz_convert('Asia/Kolkata')
                    else:
                        df_tech.index = df_tech.index.tz_convert('America/New_York')

                    # 3. Calculate ALL Indicators mathematically first
                    df_tech.ta.sma(length=20, append=True) 
                    df_tech.ta.ema(length=50, append=True) 
                    df_tech.ta.bbands(length=20, std=2, append=True)
                    df_tech.ta.rsi(length=14, append=True)
                    df_tech.ta.macd(fast=12, slow=26, signal=9, append=True)
                    
                    if 'Volume' in df_tech.columns and not df_tech['Volume'].eq(0).all():
                        df_tech.ta.vwap(append=True)
                        has_vwap = True
                    else:
                        has_vwap = False

                    df_tech.dropna(inplace=True)

                    # Create color array for Volume bars (Green for up, Red for down)
                    colors = ['#10b981' if row['Close'] >= row['Open'] else '#ef4444' for index, row in df_tech.iterrows()]

                    # 4. Build the Plotly Multi-Pane Chart (NOW 4 ROWS FOR VOLUME)
                    fig_tech = make_subplots(rows=4, cols=1, shared_xaxes=True, 
                                        vertical_spacing=0.02, 
                                        row_heights=[0.5, 0.15, 0.15, 0.2], # Optimized heights
                                        subplot_titles=("Price Action & Overlays", "Volume", "MACD Momentum", "RSI (14)"))

                    # --- ROW 1: Price & Overlays ---
                    fig_tech.add_trace(go.Candlestick(x=df_tech.index, open=df_tech['Open'], high=df_tech['High'], low=df_tech['Low'], close=df_tech['Close'], name="Price"), row=1, col=1)
                    
                    sma_col = df_tech.columns[df_tech.columns.str.contains('SMA_')][0]
                    ema_col = df_tech.columns[df_tech.columns.str.contains('EMA_')][0]
                    fig_tech.add_trace(go.Scatter(x=df_tech.index, y=df_tech[sma_col], line=dict(color='#facc15', width=1.5), name="20 SMA"), row=1, col=1)
                    fig_tech.add_trace(go.Scatter(x=df_tech.index, y=df_tech[ema_col], line=dict(color='#3b82f6', width=1.5), name="50 EMA"), row=1, col=1)
                    
                    if has_vwap:
                        vwap_col = df_tech.columns[df_tech.columns.str.startswith('VWAP')][0]
                        fig_tech.add_trace(go.Scatter(x=df_tech.index, y=df_tech[vwap_col], line=dict(color='#ec4899', width=2, dash='dot'), name="VWAP"), row=1, col=1)

                    bb_upper = df_tech.columns[df_tech.columns.str.contains('BBU_')][0]
                    bb_lower = df_tech.columns[df_tech.columns.str.contains('BBL_')][0]
                    fig_tech.add_trace(go.Scatter(x=df_tech.index, y=df_tech[bb_upper], line=dict(color='gray', width=1, dash='dot'), name="BB Upper"), row=1, col=1)
                    fig_tech.add_trace(go.Scatter(x=df_tech.index, y=df_tech[bb_lower], line=dict(color='gray', width=1, dash='dot'), name="BB Lower", fill='tonexty', fillcolor='rgba(128,128,128,0.05)'), row=1, col=1)

                    # --- ROW 2: VOLUME (NEW) ---
                    fig_tech.add_trace(go.Bar(x=df_tech.index, y=df_tech['Volume'], marker_color=colors, name="Volume"), row=2, col=1)

                    # --- ROW 3: MACD ---
                    macd_line = df_tech.columns[df_tech.columns.str.startswith('MACD_')][0]
                    macd_signal = df_tech.columns[df_tech.columns.str.startswith('MACDs_')][0]
                    macd_hist = df_tech.columns[df_tech.columns.str.startswith('MACDh_')][0]
                    
                    fig_tech.add_trace(go.Bar(x=df_tech.index, y=df_tech[macd_hist], marker_color=np.where(df_tech[macd_hist]>=0, '#10b981', '#ef4444'), name="Histogram"), row=3, col=1)
                    fig_tech.add_trace(go.Scatter(x=df_tech.index, y=df_tech[macd_line], line=dict(color='#3b82f6', width=1.5), name="MACD"), row=3, col=1)
                    fig_tech.add_trace(go.Scatter(x=df_tech.index, y=df_tech[macd_signal], line=dict(color='#f97316', width=1.5), name="Signal"), row=3, col=1)

                    # --- ROW 4: RSI ---
                    rsi_col = df_tech.columns[df_tech.columns.str.contains('RSI_')][0]
                    fig_tech.add_trace(go.Scatter(x=df_tech.index, y=df_tech[rsi_col], line=dict(color='#c084fc', width=1.5), name="RSI"), row=4, col=1)
                    fig_tech.add_hline(y=70, line_dash="dash", line_color="red", row=4, col=1)
                    fig_tech.add_hline(y=30, line_dash="dash", line_color="green", row=4, col=1)

                    # ==========================================
                    # 5. THE PERFECT ZOOM & SQUISH FIX (FOR ALL ROWS)
                    # ==========================================
                    if yf_interval == "1m": zoom_candles = 60
                    elif yf_interval == "5m": zoom_candles = 50
                    elif yf_interval == "15m": zoom_candles = 40
                    elif yf_interval == "30m": zoom_candles = 30
                    else: zoom_candles = 90
                    zoom_candles = min(zoom_candles, len(df_tech))

                    x_start, x_end = df_tech.index[-zoom_candles], df_tech.index[-1]
                    visible_data = df_tech.iloc[-zoom_candles:]

                    # Calculate precise Y-axis bounds for EVERY subplot based only on visible data
                    y_min, y_max = visible_data['Low'].min(), visible_data['High'].max()
                    y_pad = (y_max - y_min) * 0.05 
                    
                    vol_max = visible_data['Volume'].max() * 1.1 # 10% padding above volume
                    
                    macd_min = visible_data[[macd_line, macd_signal, macd_hist]].min().min()
                    macd_max = visible_data[[macd_line, macd_signal, macd_hist]].max().max()
                    macd_pad = abs(macd_max - macd_min) * 0.1

                    # Force the camera zoom on X-axis
                    fig_tech.update_xaxes(range=[x_start, x_end])
                    
                    # Force the Auto-Scale on ALL Y-axes
                    fig_tech.update_yaxes(range=[y_min - y_pad, y_max + y_pad], row=1, col=1) # Price
                    fig_tech.update_yaxes(range=[0, vol_max], row=2, col=1)                   # Volume
                    fig_tech.update_yaxes(range=[macd_min - macd_pad, macd_max + macd_pad], row=3, col=1) # MACD

                    # Hide weekend gaps
                    fig_tech.update_xaxes(rangebreaks=[dict(bounds=["sat", "mon"])])

                    fig_tech.update_layout(
                        height=900, 
                        template="plotly_dark",
                        showlegend=True,
                        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                        xaxis_rangeslider_visible=False,
                        margin=dict(l=10, r=10, t=50, b=10)
                    )
                    
                    st.plotly_chart(fig_tech, use_container_width=True)
                    # Put this right above your charting code
                    # This forces Streamlit to re-download the latest yfinance data
                    st.button("🔄 Refresh Live Data")

                    # 6. The AI Confluence Matrix
                    latest_rsi = df_tech[rsi_col].iloc[-1]
                    latest_close = df_tech['Close'].iloc[-1]
                    latest_sma = df_tech[sma_col].iloc[-1]
                    
                    st.markdown("### Technical Matrix")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        if latest_rsi > 70: st.error(f"**RSI ({latest_rsi:.1f}):** Overbought (Bearish)")
                        elif latest_rsi < 30: st.success(f"**RSI ({latest_rsi:.1f}):** Oversold (Bullish)")
                        else: st.info(f"**RSI ({latest_rsi:.1f}):** Neutral")
                            
                    with col2:
                        if latest_close > latest_sma: st.success(f"**Trend:** Bullish (Above 20 SMA)")
                        else: st.error(f"**Trend:** Bearish (Below 20 SMA)")
                            
                    with col3:
                        if has_vwap:
                            latest_vwap = df_tech[vwap_col].iloc[-1]
                            if latest_close > latest_vwap: st.success(f"**Intraday:** Bullish (Above VWAP)")
                            else: st.error(f"**Intraday:** Bearish (Below VWAP)")
                        else:
                            st.warning("**Note:** Chart is powered by custom yfinance data. AI calculates momentum underneath.")

            except Exception as e:
                st.error(f"Engine logic error: {e}")

            # --- NEW: Educational Glossary & AI Conclusion ---
            st.markdown("<br>", unsafe_allow_html=True) # Adds a little breathing room
            st.markdown("#### 📚 Indicator Glossary & Ideal Market Conditions")
            
            with st.expander("Click to view technical definitions and ideal setups"):
                st.markdown("""
                * **MACD Histogram:** Measures the distance between the MACD line and its signal line, acting as an early warning system for momentum shifts. 
                * *Ideal Setup:* A rising histogram while below the zero-line suggests a "bottom" is forming and buyers are stepping in.
                * **RSI (14) - Relative Strength Index:** A momentum oscillator that measures the speed and change of price movements on a scale of 0 to 100.
                * *Ideal Setup:* A reading below 30 indicates the asset is "Oversold" (potential buy), while above 70 indicates "Overbought" (potential sell).
                * **20-Day EMA - Exponential Moving Average:** A short-term trend line that places a higher weight on recent price data, making it react faster to news.
                * *Ideal Setup:* Used by day traders to ride momentum. As long as the price stays above this line, the short-term trend is firmly bullish.
                * **50-Day SMA - Simple Moving Average:** The institutional standard for the medium-term trend. 
                * *Ideal Setup:* When a faster moving average (like the 20 EMA) crosses *above* the 50 SMA, it creates a "Golden Cross", a highly sought-after bullish signal.
                """)
                     
            #     # except Exception as e:
            #     #     st.error(f"Could not generate Excel file: {e}")
            # ==========================================
# ==========================================
        # PART 4: EXPORT & DOWNLOADS
        # ==========================================
        st.markdown("---")
        st.markdown("### 📥 IN DEAPTH ANALYTICS")
        st.write("Download the comprehensive multi-page narrative report or the raw Excel financial model.")

        with st.spinner("Compiling Export Files..."):
            try:
                # --- 1. PDF GENERATOR ---
                def generate_deep_pdf():
                    sentiment_tag = "NEUTRAL" 
                    if "BULLISH" in nlp_text: sentiment_tag = "BULLISH"
                    elif "BEARISH" in nlp_text: sentiment_tag = "BEARISH"

                    roe_val = info.get('returnOnEquity', 0) * 100
                    try:
                        raw_equity = bs['Stockholders Equity'].iloc[-1]
                        safe_equity = 0 if pd.isna(raw_equity) else (raw_equity / 1e9)
                    except: safe_equity = 0

                    is_positive = ni.iloc[-1] > 0
                    fcf_is_positive = fcf.iloc[-1] > 0

                    health_status = "robust operational integrity" if is_positive else "significant operational challenges"
                    sustainability_msg = "The gap between cash made and what is needed shows great long-term health." if is_positive else "The gap between cash coming in and rising bills shows serious risks for the business."
                    management_msg = "management is putting cash into projects that grow the business." if is_positive else "management is struggling to keep the business profitable against high costs."
                    final_action = "buy other companies or reward shareholders with extra dividends." if is_positive else "focus on cutting debt and saving cash to keep the business running."

                    pdf = FPDF()
                    pdf.add_page()
                    pdf.set_font("Arial", 'B', 16)
                    pdf.multi_cell(0, 10, f"FINANCIAL RESEARCH BREAKDOWN: {name.upper()}", align='C')
                    
                    pdf.set_font("Arial", 'I', 10)
                    pdf.cell(0, 10, f"Market Price: {pdf_sym}{price:,.2f} | Consensus Target: {pdf_sym}{intrinsic_val:,.2f} | Base Currency: {currency_code}", ln=True, align='C')
                    pdf.line(10, 32, 200, 32)
                    pdf.ln(12)
                    
                    if isinstance(nlp_text, tuple): safe_nlp = clean_for_pdf(nlp_text[0])
                    else: safe_nlp = clean_for_pdf(nlp_text)
                    safe_consensus = clean_for_pdf(consensus_text)
                    
                    sentiment_interpretation = f"What does the sentiment mean? The NLP engine found a {sentiment_tag} mood in the news. This means the latest stories are highlighting growth chances or risks that aren't fully shown in the stock price yet."
                    financial_interpretation = f"What does the financial health mean? With a ROE of {roe_val:.2f}%, {management_msg}"
                    
                    sections = {
                        "1. Executive Summary & Intrinsic Valuation": f"This document provides a comprehensive financial breakdown of {name}. The asset is currently trading at a market price of {pdf_sym}{price:,.2f}. Based on alternative data and Wall Street consensus, the intrinsic target price is modeled at {pdf_sym}{intrinsic_val:,.2f}. Our analysis strips away standard accounting noise to focus on operational cash flow, capital structure, and true margin efficiency to determine the long-term viability of the underlying business model.",
                        "2. AI Sentiment & NLP Insights": f"{safe_nlp}",
                        "3. Consensus Forensic Breakdown": f"{safe_consensus}",
                        "4. Top-Line Trajectory & Margin Efficiency": f"Growth metrics over the observed {slice_idx}-year period reveal that total revenue successfully scaled to {pdf_sym}{rev.iloc[-1]/1e9:.2f} Billion. This top-line growth is supported by a gross margin of {(gp.iloc[-1]/rev.iloc[-1])*100:.1f}%. By managing their office and daily costs, the firm secured an operating margin of {(op_inc.iloc[-1]/rev.iloc[-1])*100:.1f}%. After all taxes, the final profit margin left for shareholders is {(ni.iloc[-1]/rev.iloc[-1])*100:.1f}%.",
                        "5. Earnings Quality Forensics": f"Real cash (Operating Cash Flow) is harder to fake than paper profits. This analysis shows {name} produced {pdf_sym}{ocf.iloc[-1]/1e9:.2f} Billion in actual cash. Since this cash amount is {'higher' if ocf.iloc[-1]>ni.iloc[-1] else 'lower'} than the reported profit of {pdf_sym}{ni.iloc[-1]/1e9:.2f} Billion, we can see how honest and strong their accounting really is.",
                        "6. Capital Reinvestment (CapEx) & Free Cash Flow": f"To keep the business running and growing, the company spent {pdf_sym}{capex.iloc[-1]/1e9:.2f} Billion on equipment and upgrades. After paying for these, the leftover 'Free Cash Flow' is {pdf_sym}{fcf.iloc[-1]/1e9:.2f} Billion.",
                        "7. Capital Structure, Liquidity & Default Risk": f"The balance sheet shows the company has {pdf_sym}{safe_equity:.2f} Billion in value for owners against debts of {pdf_sym}{info.get('totalDebt', 0)/1e9:.2f} Billion.",
                        "8. M&A Strategic Velocity & Portfolio Impact": f"The company spends about {pdf_sym}{acquisitions.mean()/1e9:.2f}B a year buying other businesses. This M&A spending makes up {(acquisitions.sum()/total_investment.sum())*100:.1f}% of their total investment.",
                        "9. Final Strategic Verdict": f"The multi-year data confirms that {name} shows {health_status}. {sustainability_msg} \n\n{sentiment_interpretation} \n\n{financial_interpretation} \n\nIn light of these metrics, management should {final_action} as a priority to secure the company's long-term future."
                    }
                    
                    for title, text in sections.items():
                        pdf.set_font("Arial", 'B', 12)
                        pdf.cell(0, 10, title, ln=True)
                        pdf.set_font("Arial", '', 11)
                        pdf.multi_cell(0, 6, text.encode('latin-1', 'replace').decode('latin-1'))
                        pdf.ln(8)
                        
                    return pdf.output(dest="S").encode("latin-1")

                # --- 2. RAW FINANCIALS & CHARTS EXCEL GENERATOR ---
                def generate_raw_excel():
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        df_formatted = pd.DataFrame({
                            'Revenue': rev.apply(lambda x: fin_sym + format_kmb(x).replace('$', '')),
                            'Net Income': ni.apply(lambda x: fin_sym + format_kmb(x).replace('$', '')),
                            'Operating Cash Flow': ocf.apply(lambda x: fin_sym + format_kmb(x).replace('$', '')),
                            'CapEx': capex.apply(lambda x: fin_sym + format_kmb(x).replace('$', '')),
                            'Free Cash Flow (UFCF)': fcf.apply(lambda x: fin_sym + format_kmb(x).replace('$', '')),
                            'Total Equity': equity.apply(lambda x: fin_sym + format_kmb(x).replace('$', '')),
                            'Total Debt': debt.apply(lambda x: fin_sym + format_kmb(x).replace('$', ''))
                        })
                        
                        df_raw = pd.DataFrame({
                            'Revenue': rev/1e9, 
                            'Gross Profit': gp/1e9,
                            'Net Income': ni/1e9,
                            'Op Cash Flow': ocf/1e9,
                            'CapEx': capex/1e9,
                            'Free Cash Flow': fcf/1e9,
                            'Total Equity': equity/1e9,
                            'Total Debt': debt/1e9
                        })
                        
                        df_formatted.to_excel(writer, sheet_name='Financial Model')
                        df_raw.to_excel(writer, sheet_name='Chart Data')
                        
                        workbook = writer.book
                        worksheet = writer.sheets['Financial Model']
                        max_row = len(years) + 1

                        chart1 = workbook.add_chart({'type': 'column'})
                        chart1.add_series({'name': "='Chart Data'!$B$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$B$2:$B${max_row}", 'fill': {'color': '#3b82f6'}})
                        chart1.add_series({'name': "='Chart Data'!$D$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$D$2:$D${max_row}", 'fill': {'color': '#10b981'}})
                        chart1.set_title({'name': f'Revenue vs Net Income ({fin_sym} Billions)'})
                        worksheet.insert_chart('J2', chart1)

                        chart2 = workbook.add_chart({'type': 'column'})
                        chart2.add_series({'name': "='Chart Data'!$E$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$E$2:$E${max_row}", 'fill': {'color': '#14b8a6'}})
                        chart2.add_series({'name': "='Chart Data'!$F$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$F$2:$F${max_row}", 'fill': {'color': '#ef4444'}})
                        chart2.add_series({'name': "='Chart Data'!$G$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$G$2:$G${max_row}", 'type': 'line', 'line': {'color': '#10b981', 'width': 3}})
                        chart2.set_title({'name': f'Cash Flow Engine ({fin_sym} Billions)'})
                        worksheet.insert_chart('J17', chart2)

                        chart3 = workbook.add_chart({'type': 'column'})
                        chart3.add_series({'name': "='Chart Data'!$H$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$H$2:$H${max_row}", 'fill': {'color': '#10b981'}})
                        chart3.add_series({'name': "='Chart Data'!$I$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$I$2:$I${max_row}", 'fill': {'color': '#ef4444'}})
                        chart3.set_title({'name': f'Capital Structure ({fin_sym} Billions)'})
                        worksheet.insert_chart('R2', chart3)

                        chart4 = workbook.add_chart({'type': 'column'})
                        chart4.add_series({'name': "='Chart Data'!$B$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$B$2:$B${max_row}", 'fill': {'color': '#3b82f6'}})
                        chart4.add_series({'name': "='Chart Data'!$C$1", 'categories': f"='Chart Data'!$A$2:$A${max_row}", 'values': f"='Chart Data'!$C$2:$C${max_row}", 'fill': {'color': '#f59e0b'}})
                        chart4.set_title({'name': f'Revenue vs Gross Profit ({fin_sym} Billions)'})
                        worksheet.insert_chart('R17', chart4)
                        
                        df_def = pd.DataFrame({
                            'Metric': ['Revenue', 'CapEx', 'Free Cash Flow (UFCF)', 'Current Ratio', 'Debt-to-Equity'],
                            'Analyst Definition': [
                                'Total top-line sales generated by the core business.',
                                'Capital Expenditures: Money spent on buying or upgrading physical assets.',
                                'Unlevered Free Cash Flow: Pure cash left over after expenses and CapEx.',
                                'Short-term liquidity. Current Assets divided by Current Liabilities.',
                                'Total Debt divided by Shareholder Equity. Measures long-term risk.'
                            ]
                        })
                        df_def.to_excel(writer, sheet_name='Metric Guide', index=False)
                        
                    return output.getvalue()

                # --- 3. DCF VARIABLES SETUP ---
                ticker_obj = yf.Ticker(ticker) 
                info = ticker_obj.info
                financials = ticker_obj.financials
                balance_sheet = ticker_obj.balance_sheet
                current_symbol = info.get('symbol', ticker).upper()
                
                def hunt_for_metric(df, aliases):
                    if df is not None and not df.empty:
                        for alias in aliases:
                            if alias in df.index:
                                val = df.loc[alias].iloc[0]
                                if pd.notna(val) and val != 0: return val
                    return None

                rev_aliases = ['Total Revenue', 'Operating Revenue', 'Revenue', 'Net Sales', 'Sales/Revenue', 'Total Operating IncomeAsReported']
                current_revenue = hunt_for_metric(financials, rev_aliases)
                if not current_revenue: current_revenue = info.get('totalRevenue', 1000000000)

                debt_aliases = ['Total Debt', 'Long Term Debt', 'Short Long Term Debt', 'Net Debt', 'Long Term Debt And Capital Lease Obligation']
                total_debt = hunt_for_metric(balance_sheet, debt_aliases)
                if not total_debt: total_debt = info.get('totalDebt', 0)

                cash_aliases = ['Cash And Cash Equivalents', 'Cash Cash Equivalents And Short Term Investments', 'Total Cash']
                total_cash = hunt_for_metric(balance_sheet, cash_aliases)
                if not total_cash: total_cash = info.get('totalCash', 0)

                shares_out = info.get('sharesOutstanding')
                if not shares_out or pd.isna(shares_out) or shares_out == 0: shares_out = info.get('impliedSharesOutstanding', 1000000)

                actual_margin = info.get('operatingMargins')
                if not actual_margin or pd.isna(actual_margin): actual_margin = 0.08
                
                actual_growth = info.get('revenueGrowth')
                if not actual_growth or pd.isna(actual_growth): actual_growth = 0.05

                stock_currency = info.get('currency', 'USD')
                financial_currency = info.get('financialCurrency', 'USD')
                
                live_fx_rate = 1.0
                if stock_currency == "INR" and financial_currency == "USD":
                    try:
                        fx_info = yf.Ticker("USDINR=X").info
                        live_fx_rate = fx_info.get('regularMarketPrice') or fx_info.get('previousClose') or 83.5
                    except Exception: live_fx_rate = 83.5

                if not current_revenue or pd.isna(current_revenue) or current_revenue == 0: current_revenue = 1000000000 
                if not total_debt or pd.isna(total_debt): total_debt = 0
                if not total_cash or pd.isna(total_cash): total_cash = 0

                # --- 4. COMPILE DCF EXCEL ---
                excel_dcf_data = generate_dcf_excel(
                    ticker_symbol=current_symbol, 
                    base_revenue=current_revenue, 
                    shares_out=shares_out, 
                    total_cash=total_cash, 
                    total_debt=total_debt,
                    actual_margin=actual_margin,
                    actual_growth=actual_growth,
                    stock_currency=stock_currency,       
                    financial_currency=financial_currency, 
                    live_fx_rate=live_fx_rate            
                )
                # ==========================================
                # 5. RENDER THE 3 DOWNLOAD BUTTONS (NO-RELOAD HTML BYPASS)
                # ==========================================
                st.markdown("<br>", unsafe_allow_html=True)
                c_exp1, c_exp2, c_exp3 = st.columns(3)
                
                # 1. PDF Generator
                pdf_data = generate_deep_pdf()
                b64_pdf = base64.b64encode(pdf_data).decode()
                href_pdf = f'<a href="data:application/pdf;base64,{b64_pdf}" download="{ticker}_Research.pdf" style="display: block; text-align: center; padding: 0.5em; background-color: #1e293b; color: white; text-decoration: none; border-radius: 5px; border: 1px solid #334155; font-family: sans-serif;">📄 Download PDF Narrative Report</a>'
                c_exp1.markdown(href_pdf, unsafe_allow_html=True)

                # 2. DCF Model Generator
                b64_dcf = base64.b64encode(excel_dcf_data.getvalue()).decode()
                href_dcf = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_dcf}" download="{current_symbol}_Automated_DCF.xlsx" style="display: block; text-align: center; padding: 0.5em; background-color: #1e293b; color: white; text-decoration: none; border-radius: 5px; border: 1px solid #334155; font-family: sans-serif;">📊 Download Proprietary DCF Model</a>'
                c_exp2.markdown(href_dcf, unsafe_allow_html=True)

                # 3. Raw Financials Generator
                raw_excel_data = generate_raw_excel()
                b64_raw = base64.b64encode(raw_excel_data).decode()
                href_raw = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_raw}" download="{ticker}_Data_Dump.xlsx" style="display: block; text-align: center; padding: 0.5em; background-color: #1e293b; color: white; text-decoration: none; border-radius: 5px; border: 1px solid #334155; font-family: sans-serif;">📈 Download Raw Financials & Charts</a>'
                c_exp3.markdown(href_raw, unsafe_allow_html=True)
            except Exception as e:
                st.error(f"Could not generate export files: {e}")
                
            # ==========================================
            # PART 3: THE MACHINE LEARNING PREDICTOR
            # ==========================================
            st.markdown("---")
            st.markdown("### Trend Prediction based on ML")

            # 1. Create a safe, permanent box for our loading message
            loading_box = st.empty()
            loading_box.info("⏳ Training Random Forest model in the background. Please wait...")

            # 2. Set up variables
            prediction, confidence, accuracy, ml_error = None, None, None, None

            # 3. The Brain: Heavy math (ABSOLUTELY NO UI COMMANDS IN HERE)
            try:
                from sklearn.ensemble import RandomForestClassifier
                import numpy as np

                df_ml = df_tech.copy()
                df_ml['Target'] = np.where(df_ml['Close'].shift(-1) > df_ml['Close'], 1, 0)
                
                # Make sure these variables match what you named them in your technical indicators!
                features = [sma_col, ema_col, rsi_col, macd_line, macd_hist] 
                df_ml = df_ml.dropna()

                X = df_ml[features]
                y = df_ml['Target']

                X_train = X.iloc[:-1]
                y_train = y.iloc[:-1]
                X_live = X.iloc[[-1]]

                rf_model = RandomForestClassifier(n_estimators=100, random_state=42, max_depth=5)
                rf_model.fit(X_train, y_train)

                prediction = rf_model.predict(X_live)[0]
                confidence = rf_model.predict_proba(X_live)[0].max() * 100
                accuracy = rf_model.score(X_train, y_train) * 100

            except Exception as e:
                ml_error = e

            # 4. Delete the loading message safely
            loading_box.empty()

            # 5. The Face: Draw the results and Explainable AI Chart
            if ml_error:
                st.warning(f"ML Engine requires more historical data to train safely. (Error: {ml_error})")
            elif prediction is not None:
                ml1, ml2, ml3 = st.columns(3)
                
                if prediction == 1:
                    ml1.success(f"**Forecast:** BULLISH (Upward)")
                else:
                    ml1.error(f"**Forecast:** BEARISH (Downward)")
                    
                ml2.metric("AI Confidence Level", f"{confidence:.1f}%")
                ml3.metric("Model Backtest Accuracy", f"{accuracy:.1f}%")
                
                # --- NEW: Explainable AI (Feature Importance Chart) ---
                st.markdown("#### 🔍 What is driving this prediction?")
                import plotly.express as px
                import pandas as pd
                
                # Extract the AI's internal weighting
                importances = rf_model.feature_importances_
                # Clean names for the chart
                feature_names = ['50-Day SMA', '20-Day EMA', 'RSI (14)', 'MACD Line', 'MACD Histogram']
                
                importance_df = pd.DataFrame({
                    'Indicator': feature_names,
                    'Importance': importances
                }).sort_values(by='Importance', ascending=True)
                
                # Draw the Plotly Bar Chart
                fig_importance = px.bar(
                    importance_df, 
                    x='Importance', 
                    y='Indicator', 
                    orientation='h',
                    color='Importance',
                    color_continuous_scale='Blues'
                )
                
                fig_importance.update_layout(
                    margin=dict(l=20, r=20, t=20, b=20),
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    coloraxis_showscale=False,
                    height=250
                )
                
                st.plotly_chart(fig_importance, use_container_width=True)



